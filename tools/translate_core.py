"""Pure (headless-testable) core for the Dracula Avontuur translation tool.

Everything the GUI needs that does *not* touch tkinter lives here so it can be
unit-tested without a display:

* :func:`collect_rows` — gather every translatable game string (messages, verb /
  direction input words, and per-room static text) into a list of row dicts.
* :func:`analyze_message_rooms` — a best-effort static analysis of the engine
  source that attributes each message id to the room(s) it can appear in.
* CSV (and optional .xlsx) import/export that round-trips losslessly.

Row schema (all values are ``str`` so CSV round-trips exactly)::

    {
        "id":        "msg:214" | "verb:GA" | "dir:N" | "room:14" | "ui:OK_TAKE"
                     | "answer:no" | "secret:word" | "intro:1",
        "type":      "message" | "verb" | "room-text" | "ui" | "answer" | "secret" | "intro",
        "room":      "14" | "7-11" | "2, 14, 29" | "any",
        "room_name": "Zolder (herberg)" | "Donker woud / Dorpsstraat" | "",
        "dutch":     "<original source text (language nl)>",
        # then one column per target language, e.g. "en": "<translation>"
    }

The *source* language of the ``dutch`` column is always **nl** (Dutch). Target
languages are added as extra columns; they start empty and are filled by a
translator (in the GUI or in Excel after export).
"""
from __future__ import annotations

import ast
import csv
import io
from pathlib import Path

# Column order that is not a language column.
FIXED_COLUMNS = ["id", "type", "room", "room_name", "object", "dutch"]

# The engine files scanned for message->room attribution.
_ENGINE_FILES = [
    "verb_events.py",
    "room_events.py",
    "navigation.py",
    "end_of_turn.py",
    "game.py",
]

SOURCE_LANG = "nl"


# --------------------------------------------------------------------------- #
#  Static analysis: which room(s) can a given message id appear in?
# --------------------------------------------------------------------------- #
#
# Heuristic (documented, best-effort — not a proof):
#
#   The engine emits a message with one of a few call shapes, all taking the
#   world.messages *index* as a constant integer argument:
#       pr(idx)                     (room_events.py / end_of_turn.py closures)
#       pr(eng, idx)                (verb_events.py / navigation.py)
#       ...message_text(idx)        (game.py)
#       pr(50 if cond else 51)      (conditional two-message pick)
#
#   For each such call site we look at the chain of enclosing ``if`` tests and
#   take the *innermost* one that pins the current room, recognising:
#       room == N           self.room == N      eng.room == N
#       room in (a, b, ..)  (and self./eng. variants)
#       LO <= room <= HI    (chained comparison, e.g. 7 <= room <= 11)
#   Negative / open guards (``room != N``, ``room < N``) and dynamic guards
#   (``room == loc(eng, X)``, ``room == flag("dde")``) pin no static room.
#
#   A message is attributed to the UNION of the room sets of all its call sites,
#   UNLESS any site has no room guard at all — then it can appear anywhere and is
#   marked "any". Room static text is attributed to its own room directly (that
#   part is exact, not heuristic).
#
# This correctly yields e.g. msg 214 -> room 14, msg 45 -> rooms 7-11, while
# generic verb replies and the parser-failure pool come out as "any".


_ROOM_NAMES = {"room"}  # bare Name nodes that mean "the current room"


def _is_room_ref(node: ast.AST) -> bool:
    """True if ``node`` refers to the current room (room / eng.room / self.room)."""
    if isinstance(node, ast.Name) and node.id in _ROOM_NAMES:
        return True
    if isinstance(node, ast.Attribute) and node.attr == "room":
        val = node.value
        return isinstance(val, ast.Name) and val.id in ("eng", "self")
    return False


def _const_int(node: ast.AST):
    if isinstance(node, ast.Constant) and isinstance(node.value, int) \
            and not isinstance(node.value, bool):
        return node.value
    return None


def _rooms_from_test(node: ast.AST) -> set[int]:
    """Extract the set of statically-pinned room ids from a boolean/compare test.

    Returns an empty set when the test pins no concrete room (dynamic/negative).
    """
    rooms: set[int] = set()
    if isinstance(node, ast.BoolOp):
        # and/or: union the pinned rooms of every operand.
        for value in node.values:
            rooms |= _rooms_from_test(value)
        return rooms
    if isinstance(node, ast.Compare):
        # Chained "LO <= room <= HI".
        if (len(node.ops) == 2
                and all(isinstance(op, (ast.LtE, ast.Lt)) for op in node.ops)
                and _is_room_ref(node.comparators[0])):
            lo = _const_int(node.left)
            hi = _const_int(node.comparators[1])
            if lo is not None and hi is not None:
                lo_adj = lo if isinstance(node.ops[0], ast.LtE) else lo + 1
                hi_adj = hi if isinstance(node.ops[1], ast.LtE) else hi - 1
                rooms |= set(range(lo_adj, hi_adj + 1))
            return rooms
        # Simple "room == N" / "room in (..)" (either operand order).
        if len(node.ops) == 1:
            op = node.ops[0]
            left, right = node.left, node.comparators[0]
            room_side, other = (None, None)
            if _is_room_ref(left):
                room_side, other = left, right
            elif _is_room_ref(right):
                room_side, other = right, left
            if room_side is not None:
                if isinstance(op, ast.Eq):
                    v = _const_int(other)
                    if v is not None:
                        rooms.add(v)
                elif isinstance(op, ast.In) and isinstance(other, (ast.Tuple, ast.List)):
                    for elt in other.elts:
                        v = _const_int(elt)
                        if v is not None:
                            rooms.add(v)
        return rooms
    return rooms


def _msg_ids_from_call(node: ast.Call) -> list[int]:
    """Message ids emitted by a call node, or [] if it is not a message emit."""
    func = node.func
    name = None
    if isinstance(func, ast.Name):
        name = func.id
    elif isinstance(func, ast.Attribute):
        name = func.attr
    if name not in ("pr", "message_text"):
        return []
    ids: list[int] = []
    for arg in node.args:
        v = _const_int(arg)
        if v is not None:
            ids.append(v)
        elif isinstance(arg, ast.IfExp):        # pr(50 if cond else 51)
            for branch in (arg.body, arg.orelse):
                bv = _const_int(branch)
                if bv is not None:
                    ids.append(bv)
    # pr(eng, idx) contributes only idx (eng is a Name, filtered out above).
    return ids


def _walk_for_messages(node: ast.AST, room_stack: list[set[int]], sink):
    """Recursively walk ``node``; call ``sink(msg_id, room_set_or_None)`` per emit.

    ``room_stack`` is the list of pinned-room sets of the enclosing ``if`` bodies
    we are currently inside (innermost last). ``room_set_or_None`` passed to the
    sink is the innermost non-empty set, or None if the site has no room guard.
    """
    if isinstance(node, ast.If):
        test_rooms = _rooms_from_test(node.test)
        # Body inherits the enclosing stack plus this test's pinned rooms.
        for child in node.body:
            _walk_for_messages(child, room_stack + [test_rooms], sink)
        # orelse (incl. elif) does NOT get this test's rooms.
        for child in node.orelse:
            _walk_for_messages(child, room_stack, sink)
        return
    if isinstance(node, ast.Call):
        ids = _msg_ids_from_call(node)
        if ids:
            room_set = None
            for rooms in reversed(room_stack):
                if rooms:
                    room_set = set(rooms)
                    break
            for mid in ids:
                sink(mid, room_set)
        # fall through and keep walking into args (nested calls are rare but harmless)
    for child in ast.iter_child_nodes(node):
        _walk_for_messages(child, room_stack, sink)


# Curated room attribution for messages the AST walk cannot pin. It only reads literal
# `if room == N:` guards, so messages gated by an EARLY-RETURN (`if room != X: return`),
# an OBJECT'S location (`room == loc(fixture)`), a FLAG, a NAVIGATION table (blocked
# doors / named-nav), or a COMPUTED index all fall to "any" even though they only ever
# appear in one place. Each entry below was traced to its emitting handler and verified
# against DRACULA.TXT + the EXE (see the audit in docs / the object-home + flag setters).
# Genuinely generic lines (parser/verb replies, examine-a-carried-item, random pools)
# and truly DYNAMIC lines (tied to Dracula's live position: 49, 147-152, 170, 173, 174,
# 228, 258) are intentionally NOT listed and stay "any".
_MANUAL_ROOMS: dict[int, list[int]] = {
    # castle outer/inner door status (room_events #3, computed pr(k-1))
    41: [20], 43: [20], 42: [21], 44: [21],
    13: [20, 21],                        # nav blocked: castle outer door shut
    # room-30 spiral-shaft named-nav blocked (no long ladder placed)
    14: [30], 15: [30],
    # herberg-zolder dusty chest (blaas, room 14)
    19: [14], 20: [14],
    # chop wood in a tree room (hak, _TREE_ROOMS)
    29: [7, 8, 9, 10, 15, 17, 18], 30: [7, 8, 9, 10, 15, 17, 18],
    # dig-tunnel "give a direction" (graaf, rooms 4/5/39)
    33: [4, 5, 39],
    # herberg (room 12) — the innkeeper
    57: [12], 58: [12], 59: [12], 129: [12], 130: [12], 133: [12], 134: [12],
    140: [12], 141: [12], 142: [12], 143: [12], 144: [12],
    181: [12], 182: [12], 189: [12],
    190: [12], 191: [12], 192: [12], 193: [12], 194: [12],
    195: [12], 196: [12], 197: [12],     # unreachable knife-throw text (advisory)
    # herberg two-men conversation (luister, room 13)
    136: [13], 137: [13], 138: [13],
    # grave-stone rotate (duw STEE, room 39)
    65: [39], 66: [39],
    # tower chest / tower luik (kasteeltoren, room 29)
    76: [29], 168: [29],
    # moonlit-window view (bekijk RAAM, room 29)
    219: [29], 220: [29], 221: [29],
    # visor view from inside the harnas (room 54)
    225: [54], 226: [54], 227: [54],
    # balcony rope (hang TOUW, room 25)
    86: [25],
    # fatal jump (spring, balcony/tower roof)
    218: [25, 28],
    # fill bottle at the well (vul, room 16)
    97: [16], 98: [16], 99: [16], 100: [16],
    # bed / sleep in the bedroom (room 1)
    115: [1], 116: [1], 124: [1], 125: [1],
    # house window shut (GA RAAM, nav blocked)
    235: [0, 6],
    # Dracula's coffin in the graftombe (room 37)
    236: [37], 239: [37], 240: [37], 241: [37], 242: [37], 243: [37], 245: [37],
    237: [37], 238: [37], 244: [37], 246: [37],
    # endgame patrol: each line fires in exactly one room (end_of_turn Block B)
    249: [22], 250: [30], 251: [31], 252: [32], 253: [37],
    # castle bedroom door muurvast (nav blocked)
    257: [22, 24],
    # room-24 endgame stake / cross reveal
    260: [24], 264: [24],
    # room-31 sesam / secret word
    265: [31], 266: [31], 268: [31],
    # spider (kruiskamer, room 34)
    271: [34], 272: [34],
}

# Messages with NO reachable emission path in the ORIGINAL game (reverse-engineering
# verified against DRACULA.EXE): the loader still holds their text, but no scenario ever
# shows them, so they are marked distinctly ("dead") in the tool rather than "any", and
# excluded from the message-coverage guarantee (tests/unit/test_zz_coverage_guard).
#   146 = a developer "provisional ending" leftover (no [0xe34] write anywhere);
#   198 = superseded by the near-duplicate msg 150 (the author wired 150, never 199);
#    93 = a LEES message the author left un-dispatched (0x5e never written).
DEAD_MESSAGES: set[int] = {93, 146, 198}

# Message id -> the game object(s) the message is ABOUT (examine/read/use/describe it),
# e.g. msg 201 (the Spelregels note) -> the briefje (obj 35), msg 109 -> the boek
# (obj 4). Surfaced as the tool's "object" column so translators see an object's
# context even when it can appear in any room (a carried object). Traced from the
# emitting handler's noun/object gates + the message text (see the object audit).
_MANUAL_OBJECTS: dict[int, list[int]] = {
    # examine / read / use / name a carried-or-present object -- these legitimately stay
    # room "any" (a carried object goes anywhere) but the object IS the context, which
    # is the whole point of this column (e.g. 201 -> briefje, 109 -> boek).
    14: [6], 20: [4], 24: [16, 17, 18, 33], 25: [6], 26: [15], 27: [6], 29: [28],
    30: [7], 34: [6], 35: [16], 36: [33], 37: [18], 38: [14],
    76: [21], 80: [31], 81: [31, 32], 86: [1], 91: [4], 97: [16], 98: [18], 99: [17],
    100: [17], 101: [8], 102: [8], 103: [7], 104: [7, 9], 105: [7],
    108: [2], 109: [4], 110: [5], 111: [36], 112: [1], 113: [13],
    114: [20], 115: [20], 116: [20],
    128: [3], 129: [3], 131: [3], 134: [3], 139: [1], 173: [3], 175: [0],
    181: [15], 182: [15], 201: [35], 202: [35], 204: [20], 205: [20],
    213: [4], 248: [27], 261: [32], 262: [2], 276: [36], 277: [36],
    # axe / knife thrown at the innkeeper (room 12) -- the weapon is named in every line
    190: [10, 28], 191: [5], 192: [10, 28], 193: [10, 28], 194: [10, 28],
    195: [5], 196: [5], 197: [5],
    # coffins / chests (obj-location gated): closed doodskist obj37 vs opened obj38
    236: [37], 237: [38], 238: [38], 239: [37], 240: [37], 241: [38], 242: [37],
    244: [38], 245: [37], 246: [38], 279: [38], 281: [13],
    # the spider (kruisspin, obj34) and the poison bottle (obj33) that kills it
    269: [34], 270: [34], 271: [33, 34], 272: [34],
}


def objects_label(world, obj_ids: list[int]) -> str:
    """Human-readable label for a message's related object(s): their display names."""
    names: list[str] = []
    for oid in obj_ids or []:
        o = world.objects.get(oid)
        if o and o.display_name and o.display_name not in names:
            names.append(o.display_name)
    return " / ".join(names)


def analyze_message_rooms(engine_dir: str | Path) -> dict[int, set[int] | None]:
    """Map each message id -> a set of room ids, or ``None`` for "any" (global).

    A message maps to None if *any* of its call sites has no static room guard;
    otherwise to the union of the pinned rooms across all its sites. Gaps the static
    walk cannot infer are then filled from the curated ``_MANUAL_ROOMS`` table.
    """
    engine_dir = Path(engine_dir)
    per_site: dict[int, list[set[int] | None]] = {}

    def sink(mid: int, room_set: set[int] | None):
        per_site.setdefault(mid, []).append(room_set)

    for fname in _ENGINE_FILES:
        path = engine_dir / fname
        if not path.exists():
            continue
        tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
        _walk_for_messages(tree, [], sink)

    result: dict[int, set[int] | None] = {}
    for mid, sites in per_site.items():
        if any(s is None for s in sites):
            result[mid] = None            # at least one global site -> "any"
        else:
            union: set[int] = set()
            for s in sites:
                union |= s
            result[mid] = union or None
    # Fill the "any" gaps the static walk cannot infer (object-location / flag /
    # nav-table / early-return / computed-index gates) from the curated table above.
    for mid, rooms in _MANUAL_ROOMS.items():
        if not result.get(mid):
            result[mid] = set(rooms)
    return result


# --------------------------------------------------------------------------- #
#  Formatting helpers
# --------------------------------------------------------------------------- #
def rooms_to_str(rooms: set[int] | None) -> str:
    """Compact room-set label: ``None``/empty -> "any"; runs collapse to "a-b"."""
    if not rooms:
        return "any"
    nums = sorted(rooms)
    parts: list[str] = []
    start = prev = nums[0]
    for n in nums[1:] + [None]:
        if n is not None and n == prev + 1:
            prev = n
            continue
        if prev - start >= 1:
            parts.append(f"{start}-{prev}")
        else:
            parts.append(str(start))
        if n is not None:
            start = prev = n
    return ", ".join(parts)


def _rooms_name_label(world, rooms: set[int] | None, room_name_fn) -> str:
    """A descriptive name for a message row's room set."""
    if not rooms:
        return ""
    names: list[str] = []
    for r in sorted(rooms):
        nm = room_name_fn(world, r)
        if nm not in names:
            names.append(nm)
    if len(names) == 1:
        return names[0]
    if len(names) <= 3:
        return " / ".join(names)
    return "various"


# --------------------------------------------------------------------------- #
#  Row collection
# --------------------------------------------------------------------------- #
def collect_rows(world, languages=("en",), engine_dir: str | Path | None = None,
                 room_name_fn=None) -> list[dict]:
    """Gather every translatable game string as a list of row dicts.

    Includes:
      * every non-empty message (type "message"),
      * every verb token and direction word (type "verb"),
      * every real room's static description text (type "room-text"),
      * the externalised engine lexicon — generic UI strings (type "ui"), the J/N
        answer letters (type "answer"), the secret word (type "secret") and the title
        screen (type "intro"). These used to be hardcoded Dutch literals in the .py
        sources; collecting them here is what makes the game fully translatable.

    ``languages`` is the list of target language codes to add as (empty) columns.
    ``engine_dir`` defaults to the engine package next to this repo; pass it
    explicitly in tests. ``room_name_fn`` defaults to :func:`tools.room_names.room_name`.
    """
    from engine.parser import (_VERB_TABLE, _DIR_TABLE,
                                _VERB_WORDS_NL, _DIR_WORDS_NL, _NOUN_WORDS_NL)

    if room_name_fn is None:
        from tools.room_names import room_name as room_name_fn
    if engine_dir is None:
        engine_dir = Path(__file__).resolve().parents[1] / "engine"

    langs = list(languages)
    msg_rooms = analyze_message_rooms(engine_dir)

    rows: list[dict] = []

    def blank_langs() -> dict:
        return {lang: "" for lang in langs}

    # --- messages --------------------------------------------------------- #
    for mid in sorted(world.messages):
        text = world.message_text(mid)
        if not text.strip():
            continue
        rooms = msg_rooms.get(mid)
        row = {
            "id": f"msg:{mid}",
            "type": "message",
            "room": rooms_to_str(rooms),
            "room_name": _rooms_name_label(world, rooms, room_name_fn),
            "object": objects_label(world, _MANUAL_OBJECTS.get(mid, [])),
            "dutch": text,
        }
        if mid in DEAD_MESSAGES:            # never shown in the original (RE-confirmed)
            row["room"] = "dead"
            row["room_name"] = "(onbereikbaar — nooit getoond in het spel)"
        row.update(blank_langs())
        rows.append(row)

    # --- verbs + directions (player input words) -------------------------- #
    # We show the FULL Dutch word the player types (pak, betreed, noord), not the
    # cryptic parser token (PAK, BETRE, NOOR): that is what a translator localises, and
    # the engine re-derives the target-language token from the translated word (enter ->
    # ENTE), exactly as for object nouns. The id keeps the original token so a translation
    # maps back to it. Verified: each Dutch word derives its own token, so nl is unchanged.
    for token, action in _VERB_TABLE:
        row = {
            "id": f"verb:{token}",
            "type": "verb",
            "room": "any",
            "room_name": f"actie: {action}",
            "dutch": _VERB_WORDS_NL.get(token, token),
        }
        row.update(blank_langs())
        rows.append(row)
    for token, _idx in _DIR_TABLE:
        row = {
            "id": f"dir:{token}",
            "type": "verb",
            "room": "any",
            "room_name": "richting",
            "dutch": _DIR_WORDS_NL.get(token, token),
        }
        row.update(blank_langs())
        rows.append(row)

    # --- scenery / interaction nouns (door, hatch, tree, fire, ...) -------- #
    # Nouns the engine's handlers test by a fixed Dutch token (open DOOR, examine STONE,
    # dig at HOLE, take off ARMOUR) rather than via a translatable object. Translating the
    # full word (deur -> door) lets engine/i18n derive the alias the engine consults, so
    # these commands work in the target language too (see engine/parser._NOUN_WORDS_NL).
    for token, word in _NOUN_WORDS_NL.items():
        row = {
            "id": f"noun:{token}",
            "type": "noun",
            "room": "any",
            "room_name": "omgeving/voorwerp",
            "dutch": word,
        }
        row.update(blank_langs())
        rows.append(row)

    # --- room static text ------------------------------------------------- #
    for rid in sorted(world.rooms):
        room = world.rooms[rid]
        if room.is_placeholder:
            continue
        text = room.description
        if not text.strip():
            continue
        row = {
            "id": f"room:{rid}",
            "type": "room-text",
            "room": str(rid),
            "room_name": room_name_fn(world, rid),
            "dutch": text,
        }
        row.update(blank_langs())
        rows.append(row)

    # --- object display names --------------------------------------------- #
    # The names the player SEES for each object ("Er is een <name> hier.", the
    # inventory, examine texts). They come from DRACULA.TXT's object records; without
    # them e.g. "gouden halsband" / "schatkist, vol met gouden munten" would be
    # untranslatable. The player's starting room for the object is shown as context.
    for oid in sorted(world.objects):
        obj = world.objects[oid]
        if not obj.is_real or not obj.display_name.strip():
            continue
        loc = obj.location
        in_room = loc in world.rooms and not world.rooms[loc].is_placeholder
        row = {
            "id": f"obj:{oid}",
            "type": "object",
            "room": str(loc) if in_room else "-",
            "room_name": room_name_fn(world, loc) if in_room else "",
            "object": "",
            "dutch": obj.display_name,
        }
        row.update(blank_langs())
        rows.append(row)

    # --- object input nouns ---------------------------------------------- #
    # The full words the player TYPES to refer to an object; the parser matches the
    # first 4 chars. Translating "knoflook" -> "garlic" lets the engine derive the
    # target-language token (GARL) so object nouns are typeable in that language.
    from engine.data.object_nouns import load_object_nouns_nl
    for oid, words in sorted(load_object_nouns_nl().items()):
        obj = world.objects.get(oid)
        if obj is None or not obj.is_real:
            continue
        loc = obj.location
        in_room = loc in world.rooms and not world.rooms[loc].is_placeholder
        row = {
            "id": f"objnoun:{oid}",
            "type": "obj-noun",
            "room": str(loc) if in_room else "-",
            "room_name": room_name_fn(world, loc) if in_room else "",
            "object": obj.display_name,
            "dutch": ", ".join(words),
        }
        row.update(blank_langs())
        rows.append(row)

    # --- externalised engine lexicon (was hardcoded in the .py sources) ---- #
    # Everything the player sees/types that is NOT in DRACULA.TXT: the generic parser
    # replies, tester/BUG headers, title screen, serial header, the secret word and the
    # J/N answer letters. See engine/data/lexicon.py + engine/data/strings_nl.json.
    lex = world.lexicon

    def _lex_row(rid, rtype, room, label, text):
        row = {"id": rid, "type": rtype, "room": room, "room_name": label, "dutch": text}
        row.update(blank_langs())
        rows.append(row)

    for key, value in lex.all_ui().items():
        if not str(value).strip():
            continue
        _lex_row(f"ui:{key}", "ui", "any", "engine-tekst", value)
    for kind, value in lex.all_answers().items():
        # 'yes' (save/quit prompt) / 'no' (reincarnate prompt) letters — J/N -> Y/N etc.
        _lex_row(f"answer:{kind}", "answer", "any", f"antwoord: {kind}", value)
    if str(lex.secret).strip():
        _lex_row("secret:word", "secret", "31", "wachtwoord", lex.secret)
    for i, line in enumerate(lex.intro):
        if not str(line).strip():          # blank title-screen spacer lines: keep as-is
            continue
        _lex_row(f"intro:{i}", "intro", "-", "titelscherm", line)
    if str(lex.header).strip():
        _lex_row("intro:header", "intro", "-", "serienummer", lex.header)

    for r in rows:                          # only message rows carry an object relation
        r.setdefault("object", "")
    return rows


def languages_of(rows: list[dict]) -> list[str]:
    """The target-language columns present in ``rows`` (order preserved)."""
    langs: list[str] = []
    for row in rows:
        for key in row:
            if key not in FIXED_COLUMNS and key not in langs:
                langs.append(key)
    return langs


def columns_of(rows: list[dict]) -> list[str]:
    return FIXED_COLUMNS + languages_of(rows)


# "Nederlands" is the display name of the source column, whose data key is "dutch".
SOURCE_LANGUAGE = "Nederlands"
# Search across every language column (source + all targets).
ANY_LANGUAGE = "(any language)"


def language_key(language: str) -> str:
    """Map a display language name to its row-dict key ("Nederlands" -> "dutch")."""
    return "dutch" if language == SOURCE_LANGUAGE else language


def filter_rows(rows: list[dict], text: str, language: str) -> list[dict]:
    """Rows whose text contains `text` (case-insensitive SUBSTRING match). Empty `text`
    returns all rows. `language` is a display name ("Nederlands", "English", …) mapped
    to its column; ANY_LANGUAGE searches the source + every target column. The "object"
    column is ALWAYS searched too, so e.g. "halsband" finds rows related to that object."""
    if not text:
        return list(rows)
    needle = text.lower()
    if language == ANY_LANGUAGE:
        keys = ["dutch", "object", *languages_of(rows)]
    else:
        keys = [language_key(language), "object"]
    return [r for r in rows if any(needle in str(r.get(k, "")).lower() for k in keys)]


# --------------------------------------------------------------------------- #
#  CSV import / export (utf-8-sig so Excel opens it cleanly)
# --------------------------------------------------------------------------- #
def export_csv(rows: list[dict], path: str | Path) -> None:
    columns = columns_of(rows)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({c: row.get(c, "") for c in columns})


def import_csv(path: str | Path) -> list[dict]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [dict(r) for r in reader]


def export_csv_string(rows: list[dict]) -> str:
    """Same as :func:`export_csv` but to a string (handy for tests)."""
    columns = columns_of(rows)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({c: row.get(c, "") for c in columns})
    return buf.getvalue()


def import_csv_string(text: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


# --------------------------------------------------------------------------- #
#  Optional .xlsx (only if openpyxl is already installed — a bonus, never a dep)
# --------------------------------------------------------------------------- #
def xlsx_available() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def export_xlsx(rows: list[dict], path: str | Path) -> None:
    import openpyxl
    columns = columns_of(rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "translations"
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c, "") for c in columns])
    wb.save(path)


def import_xlsx(path: str | Path) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(rows_iter)]
    out: list[dict] = []
    for values in rows_iter:
        row = {}
        for col, val in zip(header, values):
            row[col] = "" if val is None else str(val)
        out.append(row)
    return out
